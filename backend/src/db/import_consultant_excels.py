import os
from openpyxl import load_workbook
from sqlalchemy import text
from db.models.base_model import engine

def excel_to_db(excel_path, table_name):
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows or not rows[0]:
        print(f"文件 {excel_path} 无数据，跳过")
        return
    headers = [str(h).strip() if h is not None else f"col{i+1}" for i, h in enumerate(rows[0])]
    data_rows = [tuple("" if v is None else str(v) for v in row) for row in rows[1:] if any(row)]
    fields = ', '.join([f'"{h}" TEXT' for h in headers])
    placeholders = ', '.join([f":{h}" for h in headers])
    with engine.begin() as conn:
        # Create table if not exists
        conn.execute(text(f'CREATE TABLE IF NOT EXISTS "{table_name}" ({fields})'))
        # Insert data
        for row in data_rows:
            values_dict = {h: v for h, v in zip(headers, row)}
            conn.execute(
                text(f'INSERT INTO "{table_name}" VALUES ({", ".join([f":{h}" for h in headers])})'),
                values_dict
            )
    print(f"已导入 {excel_path} 到表 {table_name}")

def filename_to_table(filename):
    name = os.path.splitext(filename)[0].lower()
    if "consultant" in name:
        return "consultant"
    elif "resume" in name and "workex" not in name:
        return "resume"
    elif "workex" in name:
        return "workexresume"
    elif "active rating values" in name:
        return "active_rating_values"
    else:
        return name.replace("-", "_").replace(" ", "_")

def main():
    excel_dir = os.path.join(os.path.dirname(__file__), "../../../data/consultant")
    db_dir = os.path.join(os.path.dirname(__file__), "../../../db")
    os.makedirs(db_dir, exist_ok=True)
    # 删除旧表
    keep_tables = {"active_rating_values", "consultant", "resume", "workexresume"}
    with engine.begin() as conn:
        # List all tables
        result = conn.execute(text(
            "SELECT table_name FROM information_schema.tables WHERE table_schema = DATABASE()"
            if engine.url.get_backend_name().startswith("mysql")
            else "SELECT name as table_name FROM sqlite_master WHERE type='table'"
        ))
        all_tables = {row["table_name"] for row in result}
        for t in all_tables:
            if t not in keep_tables:
                conn.execute(text(f'DROP TABLE IF EXISTS "{t}"'))
                print(f"已删除旧表: {t}")
    # 导入新表
    files = [f for f in os.listdir(excel_dir) if f.endswith(".xlsx")]
    for f in files:
        excel_path = os.path.join(excel_dir, f)
        table_name = filename_to_table(f)
        excel_to_db(excel_path, table_name)
    print("全部导入完成。")

if __name__ == "__main__":
    main()
